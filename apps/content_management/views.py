from django.shortcuts import render
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from apps.content.models import Project, NewsEvent, SuccessStory, SuccessStoryGalleryImage, ProjectGalleryImage, NewsEventGalleryImage, FAQ
from django.views.generic import TemplateView
from django.db.models import Count, Sum, Q, Max, Avg, F, ExpressionWrapper, FloatField, Case, When, Value
from django.utils import timezone
import logging
import mimetypes
import os
import csv
import json
from datetime import timedelta
from apps.users.models import CustomUser
from django.db.models.functions import TruncMonth, TruncYear, TruncDay, TruncWeek
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from .forms import SuccessStoryForm, ProjectForm, NewsEventForm, FAQForm
from django.core.serializers.json import DjangoJSONEncoder
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# Get an instance of a logger
logger = logging.getLogger(__name__)


# --- Helper Functions ---
def handle_gallery_image_uploads(request, success_story_instance):
    """
    Helper function to handle multiple gallery image uploads for a success story.
    
    Args:
        request: The HTTP request object containing uploaded files
        success_story_instance: The SuccessStory instance to attach images to
    """
    if 'gallery_images' not in request.FILES:
        return
    
    files = request.FILES.getlist('gallery_images')
    
    # Get the current max order value for existing gallery images
    max_order = success_story_instance.gallery_images.aggregate(Max('order'))['order__max'] or -1
    
    for idx, uploaded_file in enumerate(files):
        # Read file data
        file_data = uploaded_file.read()
        
        # Get MIME type and filename
        mime_type, _ = mimetypes.guess_type(uploaded_file.name)
        filename = os.path.basename(uploaded_file.name)
        
        # Create gallery image entry
        SuccessStoryGalleryImage.objects.create(
            success_story=success_story_instance,
            image_blob=file_data,
            image_blob_mime=mime_type or 'application/octet-stream',
            image_blob_name=filename,
            order=max_order + idx + 1
        )


def handle_project_gallery_image_uploads(request, project_instance):
    """
    Helper function to handle multiple gallery image uploads for a project.
    
    Args:
        request: The HTTP request object containing uploaded files
        project_instance: The Project instance to attach images to
    """
    if 'project_gallery_images' not in request.FILES:
        return
    
    files = request.FILES.getlist('project_gallery_images')
    
    # Get the current max order value for existing gallery images
    max_order = project_instance.gallery_images.aggregate(Max('order'))['order__max'] or -1
    
    for idx, uploaded_file in enumerate(files):
        # Read file data
        file_data = uploaded_file.read()
        
        # Get MIME type and filename
        mime_type, _ = mimetypes.guess_type(uploaded_file.name)
        filename = os.path.basename(uploaded_file.name)
        
        # Create gallery image entry
        ProjectGalleryImage.objects.create(
            project=project_instance,
            image_blob=file_data,
            image_blob_mime=mime_type or 'application/octet-stream',
            image_blob_name=filename,
            order=max_order + idx + 1
        )


def handle_news_event_gallery_image_uploads(request, news_event_instance):
    """
    Helper function to handle multiple gallery image uploads for a news/event.
    
    Args:
        request: The HTTP request object containing uploaded files
        news_event_instance: The NewsEvent instance to attach images to
    """
    if 'news_event_gallery_images' not in request.FILES:
        return
    
    files = request.FILES.getlist('news_event_gallery_images')
    
    # Get the current max order value for existing gallery images
    max_order = news_event_instance.gallery_images.aggregate(Max('order'))['order__max'] or -1
    
    for idx, uploaded_file in enumerate(files):
        # Read file data
        file_data = uploaded_file.read()
        
        # Get MIME type and filename
        mime_type, _ = mimetypes.guess_type(uploaded_file.name)
        filename = os.path.basename(uploaded_file.name)
        
        # Create gallery image entry
        NewsEventGalleryImage.objects.create(
            news_event=news_event_instance,
            image_blob=file_data,
            image_blob_mime=mime_type or 'application/octet-stream',
            image_blob_name=filename,
            order=max_order + idx + 1
        )


# --- Dashboard Views --- 
class DashboardDataView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    JSON API view to serve aggregated data for dashboard charts.
    Accepts GET parameters:
    - range: '30d', '90d', '6m', '1y', 'all'
    - frequency: 'day', 'week', 'month'
    """
    def test_func(self):
        return self.request.user.is_staff

    def get(self, request, *args, **kwargs):
        range_param = request.GET.get('range', '6m')
        freq_param = request.GET.get('frequency', 'month')
        
        # Calculate start date
        today = timezone.now()
        start_date = today - timedelta(days=180) # Default 6m
        
        if range_param == '30d':
            start_date = today - timedelta(days=30)
        elif range_param == '90d':
            start_date = today - timedelta(days=90)
        elif range_param == '6m':
            start_date = today - timedelta(days=180)
        elif range_param == '1y':
            start_date = today - timedelta(days=365)
        elif range_param == 'all':
            start_date = today - timedelta(days=365*5) # Cap at 5 years for 'all' to avoid crazy loads
        
        # Determine Trunc function
        if freq_param == 'day':
            trunc_func = TruncDay
        elif freq_param == 'week':
            trunc_func = TruncWeek
        else:
            trunc_func = TruncMonth
            
        # 1. User Growth Over Time
        user_growth = (
            CustomUser.objects
            .filter(date_joined__gte=start_date)
            .annotate(period=trunc_func('date_joined'))
            .values('period')
            .annotate(count=Count('id'))
            .order_by('period')
        )
        
        user_labels = [item['period'].strftime('%Y-%m-%d') for item in user_growth]
        user_data = [item['count'] for item in user_growth]
        
        # 2. Content Creation Activity (Stacked)
        # We need to normalize dates across all 3 content types
        projects = (
            Project.objects
            .filter(created_at__gte=start_date)
            .annotate(period=trunc_func('created_at'))
            .values('period')
            .annotate(count=Count('id'))
            .order_by('period')
        )
        
        news = (
            NewsEvent.objects
            .filter(publish_date__gte=start_date)
            .annotate(period=trunc_func('publish_date'))
            .values('period')
            .annotate(count=Count('id'))
            .order_by('period')
        )
        
        stories = (
            SuccessStory.objects
            .filter(published_at__gte=start_date)
            .annotate(period=trunc_func('published_at'))
            .values('period')
            .annotate(count=Count('id'))
            .order_by('period')
        )
        
        # Create a set of all periods found
        all_periods = set()
        for qs in [projects, news, stories]:
            for item in qs:
                all_periods.add(item['period'])
        
        sorted_periods = sorted(list(all_periods))
        content_labels = [p.strftime('%Y-%m-%d') for p in sorted_periods]
        
        # Map data to periods
        def map_data(qs_data, periods):
            data_map = {item['period']: item['count'] for item in qs_data}
            return [data_map.get(p, 0) for p in periods]
            
        project_data = map_data(projects, sorted_periods)
        news_data = map_data(news, sorted_periods)
        story_data = map_data(stories, sorted_periods)
        
        # 3. Projects by Theme (Pie/Doughnut)
        themes = (
            Project.objects
            .filter(created_at__gte=start_date)
            .values('theme')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        theme_labels = [item['theme'] for item in themes]
        theme_data = [item['count'] for item in themes]
        
        # 4. Projects by Country (Bar)
        countries = (
            Project.objects
            .filter(created_at__gte=start_date)
            .values('country')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        country_labels = [item['country'] for item in countries]
        country_data = [item['count'] for item in countries]

        # 5. Beneficiaries Impact (Top 5 Projects created in range)
        top_projects_impact = (
             Project.objects
             .filter(created_at__gte=start_date, headcount__gt=0)
             .order_by('-headcount')[:5]
             .values('title', 'headcount')
        )
        impact_labels = [item['title'][:20] + '...' for item in top_projects_impact]
        impact_data = [item['headcount'] for item in top_projects_impact]

        # 6. User Demographics (Gender)
        # Note: Not filtered by date range as demographics are usually "current state"
        gender_dist = (
            CustomUser.objects
            .exclude(gender='')
            .values('gender')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        gender_labels = [item['gender'] for item in gender_dist]
        gender_data = [item['count'] for item in gender_dist]

        # 7. Projects by Difficulty
        # Filtered by date range to show mix of recent projects
        difficulty_dist = (
            Project.objects
            .filter(created_at__gte=start_date)
            .values('difficulty')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        difficulty_labels = [item['difficulty'] for item in difficulty_dist]
        difficulty_data = [item['count'] for item in difficulty_dist]

        # 8. User Countries (Top 5)
        user_countries = (
            CustomUser.objects
            .exclude(country_code='')
            .values('country_code')
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
        )
        user_country_labels = [item['country_code'] for item in user_countries]
        user_country_data = [item['count'] for item in user_countries]

        # 9. Login Methods
        login_methods = (
            CustomUser.objects
            .values('login_method')
            .annotate(count=Count('id'))
            .order_by('-count')
        )
        login_method_labels = [item['login_method'] for item in login_methods]
        login_method_data = [item['count'] for item in login_methods]
        
        data = {
            'user_growth': {
                'labels': user_labels,
                'data': user_data
            },
            'content_activity': {
                'labels': content_labels,
                'datasets': {
                    'projects': project_data,
                    'news': news_data,
                    'stories': story_data
                }
            },
            'themes': {
                'labels': theme_labels,
                'data': theme_data
            },
            'countries': {
                'labels': country_labels,
                'data': country_data
            },
            'impact': {
                'labels': impact_labels,
                'data': impact_data
            },
            'demographics': {
                'gender': {
                    'labels': gender_labels,
                    'data': gender_data
                },
                'countries': {
                    'labels': user_country_labels,
                    'data': user_country_data
                },
                'login_methods': {
                    'labels': login_method_labels,
                    'data': login_method_data
                }
            },
            'project_difficulty': {
                'labels': difficulty_labels,
                'data': difficulty_data
            }
        }
        
        return JsonResponse(data)

class ManagementDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'content_management/management_dashboard.html'
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # --- User Stats ---
        total_users = CustomUser.objects.count()
        new_users_30d = CustomUser.objects.filter(date_joined__gte=timezone.now() - timezone.timedelta(days=30)).count()
        active_users_30d = CustomUser.objects.filter(last_login__gte=timezone.now() - timezone.timedelta(days=30)).count()
        
        context['total_users'] = total_users
        context['new_users_30d'] = new_users_30d
        context['active_users_30d'] = active_users_30d
        # Avoid division by zero
        previous_users = total_users - new_users_30d
        context['user_growth_pct'] = round((new_users_30d / previous_users * 100), 1) if previous_users > 0 else 100

        # User Demographics
        context['users_by_gender'] = CustomUser.objects.values('gender').annotate(count=Count('id')).exclude(gender='').order_by('-count')
        context['users_by_country'] = CustomUser.objects.values('country_code').annotate(count=Count('id')).exclude(country_code='').order_by('-count')
        context['login_methods'] = CustomUser.objects.values('login_method').annotate(count=Count('id')).order_by('-count')
        context['onboarding_complete'] = CustomUser.objects.filter(onboarding_complete=True).count()
        context['email_verified'] = CustomUser.objects.filter(email_verified=True).count()
        context['onboarding_rate'] = round((context['onboarding_complete'] / total_users * 100), 1) if total_users > 0 else 0
        context['verification_rate'] = round((context['email_verified'] / total_users * 100), 1) if total_users > 0 else 0

        # --- Project Stats ---
        context['total_projects'] = Project.objects.count()
        context['active_projects'] = Project.objects.filter(is_active=True).count()
        context['completed_projects'] = Project.objects.filter(end_date__lt=timezone.now().date()).count()
        context['upcoming_projects'] = Project.objects.filter(start_date__gt=timezone.now().date()).count()
        context['featured_projects'] = Project.objects.filter(is_featured=True).count()
        context['hero_projects'] = Project.objects.filter(is_hero_highlight=True).count()
        
        # Capacity & Enrollment
        project_aggs = Project.objects.aggregate(
            total_capacity=Sum('total_headcount'),
            total_enrolled=Sum('headcount'),
            avg_duration=Avg(F('end_date') - F('start_date'))
        )
        context['total_capacity'] = project_aggs['total_capacity'] or 0
        context['total_enrolled'] = project_aggs['total_enrolled'] or 0
        context['utilization_rate'] = round((context['total_enrolled'] / context['total_capacity'] * 100), 1) if context['total_capacity'] > 0 else 0
        context['avg_project_duration'] = project_aggs['avg_duration'].days if project_aggs['avg_duration'] else 0

        # Unique enrolled users
        context['unique_enrolled_users'] = Project.objects.values('enrolled_users').distinct().count()

        # Upcoming deadlines
        context['upcoming_deadlines'] = Project.objects.filter(
            application_deadline__gte=timezone.now(),
            application_deadline__lte=timezone.now() + timezone.timedelta(days=30)
        ).count()

        # Breakdowns
        context['projects_by_difficulty'] = Project.objects.values('difficulty').annotate(count=Count('id')).order_by('-count')
        context['projects_by_theme'] = Project.objects.values('theme').annotate(count=Count('id')).order_by('-count')
        context['projects_by_country'] = Project.objects.values('country').annotate(count=Count('id')).order_by('-count')

        # --- News & Events ---
        context['total_news_events'] = NewsEvent.objects.count()
        context['published_news_events'] = NewsEvent.objects.filter(is_published=True).count()
        context['news_count'] = NewsEvent.objects.filter(content_type='News').count()
        context['event_count'] = NewsEvent.objects.filter(content_type='Event').count()
        context['featured_news'] = NewsEvent.objects.filter(is_featured=True).count()
        context['hero_news'] = NewsEvent.objects.filter(is_hero_highlight=True).count()

        # --- Success Stories & Impact ---
        context['total_success_stories'] = SuccessStory.objects.count()
        context['published_stories'] = SuccessStory.objects.filter(is_published=True).count()
        context['featured_stories'] = SuccessStory.objects.filter(is_featured=True).count()
        context['hero_stories'] = SuccessStory.objects.filter(is_hero_highlight=True).count()
        
        impact_aggs = SuccessStory.objects.aggregate(
            total_beneficiaries=Sum('beneficiaries'),
            total_hours=Sum('total_hours_contributed'),
            avg_beneficiaries=Avg('beneficiaries'),
            avg_hours=Avg('total_hours_contributed')
        )
        context['total_beneficiaries'] = impact_aggs['total_beneficiaries'] or 0
        context['total_hours_contributed'] = impact_aggs['total_hours'] or 0
        context['avg_beneficiaries_per_story'] = round(impact_aggs['avg_beneficiaries'] or 0, 1)
        context['avg_hours_per_story'] = round(impact_aggs['avg_hours'] or 0, 1)
        
        # --- FAQs ---
        context['total_faqs'] = FAQ.objects.count()

        # --- System Health & Info ---
        context['kicc_synced_projects'] = Project.objects.filter(kicc_project_id__isnull=False).exclude(kicc_project_id='').count()
        faq_votes = FAQ.objects.aggregate(
            total_up=Sum('thumbs_up'),
            total_down=Sum('thumbs_down')
        )
        total_faq_votes = (faq_votes['total_up'] or 0) + (faq_votes['total_down'] or 0)
        context['faq_helpfulness_ratio'] = round((faq_votes['total_up'] or 0) * 100 / total_faq_votes, 1) if total_faq_votes > 0 else 100
        context['admin_user_count'] = CustomUser.objects.filter(is_staff=True).count()

        # --- Gallery Images ---
        context['total_project_gallery_images'] = ProjectGalleryImage.objects.count()
        context['total_success_story_gallery_images'] = SuccessStoryGalleryImage.objects.count()
        context['total_news_event_gallery_images'] = NewsEventGalleryImage.objects.count()
        context['total_gallery_images'] = (
            context['total_project_gallery_images'] + 
            context['total_success_story_gallery_images'] + 
            context['total_news_event_gallery_images']
        )
        
        # --- Cover Images & Videos ---
        context['total_cover_images'] = (
            Project.objects.exclude(cover_image_blob__isnull=True).exclude(cover_image_blob__exact=b'').count() +
            NewsEvent.objects.exclude(cover_image_blob__isnull=True).exclude(cover_image_blob__exact=b'').count() +
            SuccessStory.objects.exclude(cover_image_blob__isnull=True).exclude(cover_image_blob__exact=b'').count()
        )
        context['total_videos'] = Project.objects.aggregate(
            total=Count('video_urls', filter=Q(video_urls__isnull=False) & ~Q(video_urls='[]'))
        )['total'] or 0

        # --- Image & Cover URLs ---
        projects_with_urls = Project.objects.filter(image_urls__isnull=False).exclude(image_urls='[]').values_list('image_urls', flat=True)
        news_with_urls = NewsEvent.objects.filter(image_urls__isnull=False).exclude(image_urls='[]').values_list('image_urls', flat=True)
        stories_with_urls = SuccessStory.objects.filter(image_urls__isnull=False).exclude(image_urls='[]').values_list('image_urls', flat=True)

        context['total_gallery_image_urls'] = sum(len(urls) for urls in projects_with_urls) + \
                                              sum(len(urls) for urls in news_with_urls) + \
                                              sum(len(urls) for urls in stories_with_urls)

        context['total_cover_image_urls'] = (
            Project.objects.filter(cover_image_url__isnull=False).exclude(cover_image_url__exact='').count() +
            NewsEvent.objects.filter(cover_image_url__isnull=False).exclude(cover_image_url__exact='').count() +
            SuccessStory.objects.filter(cover_image_url__isnull=False).exclude(cover_image_url__exact='').count()
        )

        # --- Additional Engagement Metrics ---
        context['avg_enrollment_per_project'] = round(context['total_enrolled'] / context['total_projects'], 1) if context['total_projects'] > 0 else 0
        context['content_completion_rate'] = round((context['published_stories'] + context['published_news_events']) * 100 / (context['total_success_stories'] + context['total_news_events']), 1) if (context['total_success_stories'] + context['total_news_events']) > 0 else 0

        # --- Additional Media Metrics ---
        context['total_media_items'] = (
            context['total_gallery_images'] + context['total_cover_images'] + 
            context['total_gallery_image_urls'] + context['total_cover_image_urls'] + 
            context['total_videos']
        )

        # --- Additional System Health Metrics ---
        context['total_database_records'] = (
            context['total_users'] + context['total_projects'] + 
            context['total_news_events'] + context['total_success_stories'] + 
            context['total_faqs'] + context['total_gallery_images']
        )
        context['data_integrity_score'] = round(
            (context['kicc_synced_projects'] * 100 / context['total_projects']) if context['total_projects'] > 0 else 0, 1
        )
        context['content_freshness_days'] = (timezone.now().date() - Project.objects.aggregate(latest=Max('updated_at'))['latest'].date()).days if Project.objects.exists() else 0


        # --- Recent Content ---
        context['recent_projects'] = Project.objects.order_by('-created_at')[:5]
        context['recent_news_events'] = NewsEvent.objects.order_by('-publish_date')[:5]
        context['recent_success_stories'] = SuccessStory.objects.order_by('-published_at')[:5]
        
        # --- Chart Data Preparation ---
        # 1. Content Distribution (Pie Chart)
        context['chart_labels'] = ['Projects', 'News & Events', 'Success Stories', 'FAQs']
        context['chart_data'] = [
            context['total_projects'], 
            context['total_news_events'], 
            context['total_success_stories'], 
            context['total_faqs']
        ]

        # 2. Monthly Activity (Line Chart) - Last 6 Months
        today = timezone.now()
        months = []
        project_counts = []
        news_counts = []
        
        # Simple loop for last 6 months
        for i in range(5, -1, -1):
            # Calculate start of month
            month_date = today - timezone.timedelta(days=i*30) 
            month_start = month_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            # Calculate end of month (start of next month)
            if month_start.month == 12:
                month_end = month_start.replace(year=month_start.year+1, month=1)
            else:
                month_end = month_start.replace(month=month_start.month+1)
            
            months.append(month_start.strftime('%b'))
            project_counts.append(Project.objects.filter(created_at__gte=month_start, created_at__lt=month_end).count())
            news_counts.append(NewsEvent.objects.filter(publish_date__gte=month_start, publish_date__lt=month_end).count())

        context['activity_months'] = months
        context['activity_project_data'] = project_counts
        context['activity_news_data'] = news_counts

        return context


class UserAnalyticsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = CustomUser
    template_name = 'content_management/user_analytics.html'
    context_object_name = 'viewed_users'
    paginate_by = 20
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        queryset = CustomUser.objects.all()

        # Search functionality
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(username__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(country_code__icontains=search_query)
            )

        # Sorting functionality
        sort_by = self.request.GET.get('sort', '-date_joined')
        valid_sort_fields = {
            'username': 'username',
            '-username': '-username',
            'email': 'email',
            '-email': '-email',
            'first_name': 'first_name',
            '-first_name': '-first_name',
            'last_name': 'last_name',
            '-last_name': '-last_name',
            'country_code': 'country_code',
            '-country_code': '-country_code',
            'date_joined': 'date_joined',
            '-date_joined': '-date_joined',
            'is_active': 'is_active',
            '-is_active': '-is_active',
        }

        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-date_joined')

        return queryset

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format:
            return self.export_data(export_format)
        return super().get(request, *args, **kwargs)

    def export_data(self, format_type):
        queryset = self.get_queryset()
        
        if format_type == 'csv':
            return self.export_csv()
        elif format_type == 'tsv':
            return self.export_tsv()
        elif format_type == 'json':
            return self.export_json()
        elif format_type == 'xlsx' and XLSX_AVAILABLE:
            return self.export_xlsx()
        elif format_type == 'html':
            return self.export_html()
        else:
            return self.export_csv()

    def export_csv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="user_analytics.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date of Birth', 
            'Gender', 'Blood Group', 'Guardian Name', 'Guardian Relation', 'Address', 
            'Contact', 'Country Code', 'Login Method', 'Onboarding Complete', 
            'Email Verified', 'Is Active', 'Is Staff', 'Is Superuser', 'Date Joined', 
            'Last Login'
        ])
        
        for user in queryset:
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.date_of_birth.strftime('%Y-%m-%d') if user.date_of_birth else '',
                user.gender,
                user.blood_group,
                user.guardian_name,
                user.guardian_relation,
                user.address,
                user.contact,
                user.country_code,
                user.login_method,
                'Yes' if user.onboarding_complete else 'No',
                'Yes' if user.email_verified else 'No',
                'Yes' if user.is_active else 'No',
                'Yes' if user.is_staff else 'No',
                'Yes' if user.is_superuser else 'No',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else ''
            ])
        
        return response

    def export_tsv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/tab-separated-values')
        response['Content-Disposition'] = 'attachment; filename="user_analytics.tsv"'
        
        writer = csv.writer(response, delimiter='\t')
        writer.writerow([
            'ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date of Birth', 
            'Gender', 'Blood Group', 'Guardian Name', 'Guardian Relation', 'Address', 
            'Contact', 'Country Code', 'Login Method', 'Onboarding Complete', 
            'Email Verified', 'Is Active', 'Is Staff', 'Is Superuser', 'Date Joined', 
            'Last Login'
        ])
        
        for user in queryset:
            writer.writerow([
                user.id,
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.date_of_birth.strftime('%Y-%m-%d') if user.date_of_birth else '',
                user.gender,
                user.blood_group,
                user.guardian_name,
                user.guardian_relation,
                user.address,
                user.contact,
                user.country_code,
                user.login_method,
                'Yes' if user.onboarding_complete else 'No',
                'Yes' if user.email_verified else 'No',
                'Yes' if user.is_active else 'No',
                'Yes' if user.is_staff else 'No',
                'Yes' if user.is_superuser else 'No',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else ''
            ])
        
        return response

    def export_json(self):
        queryset = self.get_queryset()
        data = []
        
        for user in queryset:
            data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'date_of_birth': user.date_of_birth.isoformat() if user.date_of_birth else None,
                'gender': user.gender,
                'blood_group': user.blood_group,
                'guardian_name': user.guardian_name,
                'guardian_relation': user.guardian_relation,
                'address': user.address,
                'contact': user.contact,
                'country_code': user.country_code,
                'login_method': user.login_method,
                'onboarding_complete': user.onboarding_complete,
                'email_verified': user.email_verified,
                'is_active': user.is_active,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser,
                'date_joined': user.date_joined.isoformat(),
                'last_login': user.last_login.isoformat() if user.last_login else None
            })
        
        response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="user_analytics.json"'
        return response

    def export_xlsx(self):
        if not XLSX_AVAILABLE:
            return HttpResponse("XLSX export not available. Please install openpyxl.", status=400)
            
        queryset = self.get_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = "User Analytics"
        
        headers = [
            'ID', 'Username', 'Email', 'First Name', 'Last Name', 'Date of Birth', 
            'Gender', 'Blood Group', 'Guardian Name', 'Guardian Relation', 'Address', 
            'Contact', 'Country Code', 'Login Method', 'Onboarding Complete', 
            'Email Verified', 'Is Active', 'Is Staff', 'Is Superuser', 'Date Joined', 
            'Last Login'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, user in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=user.id)
            ws.cell(row=row_num, column=2, value=user.username)
            ws.cell(row=row_num, column=3, value=user.email)
            ws.cell(row=row_num, column=4, value=user.first_name)
            ws.cell(row=row_num, column=5, value=user.last_name)
            ws.cell(row=row_num, column=6, value=user.date_of_birth.strftime('%Y-%m-%d') if user.date_of_birth else '')
            ws.cell(row=row_num, column=7, value=user.gender)
            ws.cell(row=row_num, column=8, value=user.blood_group)
            ws.cell(row=row_num, column=9, value=user.guardian_name)
            ws.cell(row=row_num, column=10, value=user.guardian_relation)
            ws.cell(row=row_num, column=11, value=user.address)
            ws.cell(row=row_num, column=12, value=user.contact)
            ws.cell(row=row_num, column=13, value=user.country_code)
            ws.cell(row=row_num, column=14, value=user.login_method)
            ws.cell(row=row_num, column=15, value='Yes' if user.onboarding_complete else 'No')
            ws.cell(row=row_num, column=16, value='Yes' if user.email_verified else 'No')
            ws.cell(row=row_num, column=17, value='Yes' if user.is_active else 'No')
            ws.cell(row=row_num, column=18, value='Yes' if user.is_staff else 'No')
            ws.cell(row=row_num, column=19, value='Yes' if user.is_superuser else 'No')
            ws.cell(row=row_num, column=20, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=21, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="user_analytics.xlsx"'
        wb.save(response)
        return response

    def export_html(self):
        queryset = self.get_queryset()
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>User Analytics Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                h1 {{ color: #333; }}
            </style>
        </head>
        <body>
            <h1>User Analytics Export</h1>
            <p>Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>Total records: {queryset.count()}</p>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Username</th>
                        <th>Email</th>
                        <th>First Name</th>
                        <th>Last Name</th>
                        <th>Date of Birth</th>
                        <th>Gender</th>
                        <th>Blood Group</th>
                        <th>Guardian Name</th>
                        <th>Guardian Relation</th>
                        <th>Address</th>
                        <th>Contact</th>
                        <th>Country Code</th>
                        <th>Login Method</th>
                        <th>Onboarding Complete</th>
                        <th>Email Verified</th>
                        <th>Is Active</th>
                        <th>Is Staff</th>
                        <th>Is Superuser</th>
                        <th>Date Joined</th>
                        <th>Last Login</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for user in queryset:
            html_content += f"""
                    <tr>
                        <td>{user.id}</td>
                        <td>{user.username}</td>
                        <td>{user.email}</td>
                        <td>{user.first_name}</td>
                        <td>{user.last_name}</td>
                        <td>{user.date_of_birth.strftime('%Y-%m-%d') if user.date_of_birth else ''}</td>
                        <td>{user.gender}</td>
                        <td>{user.blood_group}</td>
                        <td>{user.guardian_name}</td>
                        <td>{user.guardian_relation}</td>
                        <td>{user.address}</td>
                        <td>{user.contact}</td>
                        <td>{user.country_code}</td>
                        <td>{user.login_method}</td>
                        <td>{'Yes' if user.onboarding_complete else 'No'}</td>
                        <td>{'Yes' if user.email_verified else 'No'}</td>
                        <td>{'Yes' if user.is_active else 'No'}</td>
                        <td>{'Yes' if user.is_staff else 'No'}</td>
                        <td>{'Yes' if user.is_superuser else 'No'}</td>
                        <td>{user.date_joined.strftime('%Y-%m-%d %H:%M:%S')}</td>
                        <td>{user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else ''}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="user_analytics.html"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['current_sort'] = self.request.GET.get('sort', '-date_joined')

        # Total Users
        context['total_users'] = CustomUser.objects.count()

        # New Users Over Time (Monthly)
        context['users_by_month'] = (
            CustomUser.objects
            .annotate(month=TruncMonth('date_joined'))
            .values('month')
            .annotate(count=Count('id'))
            .order_by('month')
        )

        # New Users Over Time (Yearly)
        context['users_by_year'] = (
            CustomUser.objects
            .annotate(year=TruncYear('date_joined'))
            .values('year')
            .annotate(count=Count('id'))
            .order_by('year')
        )

        # User Demographics
        context['users_by_gender'] = CustomUser.objects.values('gender').annotate(count=Count('gender')).order_by('-count')
        context['users_by_country'] = CustomUser.objects.values('country_code').annotate(count=Count('country_code')).order_by('-count')

        # Active Users (users who have enrolled in projects)
        context['active_users'] = CustomUser.objects.filter(enrolled_projects__isnull=False).distinct().count()

        return context

class UserCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = CustomUser
    context_object_name = 'viewed_user'
    template_name = 'content_management/user_form.html'
    fields = [
        'username', 'email', 'first_name', 'last_name',
        'date_of_birth', 'gender', 'blood_group', 'guardian_name',
        'guardian_relation', 'address', 'contact', 'country_code',
        'is_active', 'is_staff', 'is_superuser'
    ]
    success_url = reverse_lazy('user_analytics')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)
    
class UserUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = CustomUser
    context_object_name = 'viewed_user'
    template_name = 'content_management/user_form.html'
    fields = [
        'username', 'email', 'first_name', 'last_name',
        'date_of_birth', 'gender', 'blood_group', 'guardian_name',
        'guardian_relation', 'address', 'contact', 'country_code',
        'is_active', 'is_staff', 'is_superuser'
    ]
    success_url = reverse_lazy('user_analytics')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)


class UserDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = CustomUser
    template_name = 'content_management/user_detail.html'
    context_object_name = 'viewed_user'
    queryset = CustomUser.objects.all()
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff


class UserDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = CustomUser
    template_name = 'content_management/user_confirm_delete.html'
    context_object_name = 'viewed_user'
    success_url = reverse_lazy('user_analytics')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff


class ApplicationAnalyticsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'content_management/application_analytics.html'
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Project Enrollment Stats
        context['total_enrollments'] = Project.objects.aggregate(total=Sum('headcount'))['total'] or 0
        context['projects_with_enrollments'] = Project.objects.filter(headcount__gt=0).count()

        # Popular Projects (by enrollment)
        context['popular_projects'] = Project.objects.filter(headcount__gt=0).order_by('-headcount')[:10]

        # Projects by Theme with enrollment counts
        context['enrollments_by_theme'] = list(
            Project.objects
            .values('theme')
            .annotate(total_enrollments=Sum('headcount'), project_count=Count('id'))
            .order_by('-total_enrollments')
        )

        # Projects by Country with enrollment counts
        context['enrollments_by_country'] = list(
            Project.objects
            .values('country')
            .annotate(total_enrollments=Sum('headcount'), project_count=Count('id'))
            .order_by('-total_enrollments')
        )

        # Project Capacity Utilization
        projects = Project.objects.filter(total_headcount__gt=0).exclude(total_headcount__isnull=True)
        context['capacity_data'] = [
            {
                'pk': p.pk,
                'title': p.title,
                'utilization': (p.headcount / p.total_headcount * 100) if p.total_headcount > 0 else 0,
                'enrolled': p.headcount,
                'capacity': p.total_headcount
            } for p in projects if p.pk
        ][:10]  # Top 10

        # Recent Enrollments (projects with recent enrollment activity)
        context['recent_enrollments'] = Project.objects.filter(headcount__gt=0).order_by('-updated_at')[:10]

        return context



# --- Project Views ---
class ProjectListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    paginate_by = 15
    model = Project
    template_name = 'content_management/project_list.html'
    context_object_name = 'projects'
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        queryset = Project.objects.all()

        # Search functionality
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) |
                Q(country__icontains=search_query) |
                Q(theme__icontains=search_query) |
                Q(teaser__icontains=search_query) |
                Q(background_objectives__icontains=search_query)
            )

        # Sorting functionality
        sort_by = self.request.GET.get('sort', '-created_at')
        valid_sort_fields = {
            'title': 'title',
            '-title': '-title',
            'country': 'country',
            '-country': '-country',
            'theme': 'theme',
            '-theme': '-theme',
            'headcount': 'headcount',
            '-headcount': '-headcount',
            'is_active': 'is_active',
            '-is_active': '-is_active',
            'created_at': 'created_at',
            '-created_at': '-created_at',
            'start_date': 'start_date',
            '-start_date': '-start_date',
        }

        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-created_at')

        return queryset

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format:
            return self.export_data(export_format)
        return super().get(request, *args, **kwargs)

    def export_data(self, format_type):
        queryset = self.get_queryset()
        
        if format_type == 'csv':
            return self.export_csv()
        elif format_type == 'tsv':
            return self.export_tsv()
        elif format_type == 'json':
            return self.export_json()
        elif format_type == 'xlsx' and XLSX_AVAILABLE:
            return self.export_xlsx()
        elif format_type == 'html':
            return self.export_html()
        else:
            # Default to CSV
            return self.export_csv()

    def export_csv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="projects.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Project ID', 'KICC Project ID', 'Title', 'Teaser', 'Background Objectives', 
            'Tasks Eligibility', 'Country', 'Theme', 'Duration', 'Difficulty', 'Headcount', 
            'Total Headcount', 'Cover Image URL', 'Video URLs', 'Image URLs', 
            'Application Deadline', 'Start Date', 'End Date', 'Is Active', 'Is Hero Highlight', 
            'Is Featured', 'Enrolled Users Count', 'Created At', 'Updated At'
        ])
        
        for project in queryset:
            writer.writerow([
                project.id,
                project.project_id,
                project.kicc_project_id or '',
                project.title,
                project.teaser,
                project.background_objectives,
                project.tasks_eligibility,
                project.country,
                project.theme,
                project.duration,
                project.difficulty,
                project.headcount or 0,
                project.total_headcount,
                project.cover_image_url or '',
                ', '.join(project.video_urls) if project.video_urls else '',
                ', '.join(project.image_urls) if project.image_urls else '',
                project.application_deadline.strftime('%Y-%m-%d %H:%M:%S') if project.application_deadline else '',
                project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
                project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
                'Yes' if project.is_active else 'No',
                'Yes' if project.is_hero_highlight else 'No',
                'Yes' if project.is_featured else 'No',
                project.enrolled_users.count(),
                project.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                project.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_tsv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/tab-separated-values')
        response['Content-Disposition'] = 'attachment; filename="projects.tsv"'
        
        writer = csv.writer(response, delimiter='\t')
        writer.writerow([
            'ID', 'Project ID', 'KICC Project ID', 'Title', 'Teaser', 'Background Objectives', 
            'Tasks Eligibility', 'Country', 'Theme', 'Duration', 'Difficulty', 'Headcount', 
            'Total Headcount', 'Cover Image URL', 'Video URLs', 'Image URLs', 
            'Application Deadline', 'Start Date', 'End Date', 'Is Active', 'Is Hero Highlight', 
            'Is Featured', 'Enrolled Users Count', 'Created At', 'Updated At'
        ])
        
        for project in queryset:
            writer.writerow([
                project.id,
                project.project_id,
                project.kicc_project_id or '',
                project.title,
                project.teaser,
                project.background_objectives,
                project.tasks_eligibility,
                project.country,
                project.theme,
                project.duration,
                project.difficulty,
                project.headcount or 0,
                project.total_headcount,
                project.cover_image_url or '',
                ', '.join(project.video_urls) if project.video_urls else '',
                ', '.join(project.image_urls) if project.image_urls else '',
                project.application_deadline.strftime('%Y-%m-%d %H:%M:%S') if project.application_deadline else '',
                project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
                project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
                'Yes' if project.is_active else 'No',
                'Yes' if project.is_hero_highlight else 'No',
                'Yes' if project.is_featured else 'No',
                project.enrolled_users.count(),
                project.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                project.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_json(self):
        queryset = self.get_queryset()
        data = []
        
        for project in queryset:
            data.append({
                'id': project.id,
                'project_id': project.project_id,
                'kicc_project_id': project.kicc_project_id,
                'title': project.title,
                'teaser': project.teaser,
                'background_objectives': project.background_objectives,
                'tasks_eligibility': project.tasks_eligibility,
                'country': project.country,
                'theme': project.theme,
                'duration': project.duration,
                'difficulty': project.difficulty,
                'headcount': project.headcount,
                'total_headcount': project.total_headcount,
                'cover_image_url': project.cover_image_url,
                'video_urls': project.video_urls,
                'image_urls': project.image_urls,
                'application_deadline': project.application_deadline.isoformat() if project.application_deadline else None,
                'start_date': project.start_date.isoformat() if project.start_date else None,
                'end_date': project.end_date.isoformat() if project.end_date else None,
                'is_active': project.is_active,
                'is_hero_highlight': project.is_hero_highlight,
                'is_featured': project.is_featured,
                'enrolled_users_count': project.enrolled_users.count(),
                'created_at': project.created_at.isoformat(),
                'updated_at': project.updated_at.isoformat()
            })
        
        response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="projects.json"'
        return response

    def export_xlsx(self):
        if not XLSX_AVAILABLE:
            return HttpResponse("XLSX export not available. Please install openpyxl.", status=400)
            
        queryset = self.get_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"
        
        # Header row with styling
        headers = [
            'ID', 'Project ID', 'KICC Project ID', 'Title', 'Teaser', 'Background Objectives', 
            'Tasks Eligibility', 'Country', 'Theme', 'Duration', 'Difficulty', 'Headcount', 
            'Total Headcount', 'Cover Image URL', 'Video URLs', 'Image URLs', 
            'Application Deadline', 'Start Date', 'End Date', 'Is Active', 'Is Hero Highlight', 
            'Is Featured', 'Enrolled Users Count', 'Created At', 'Updated At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data rows
        for row_num, project in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=project.id)
            ws.cell(row=row_num, column=2, value=project.project_id)
            ws.cell(row=row_num, column=3, value=project.kicc_project_id)
            ws.cell(row=row_num, column=4, value=project.title)
            ws.cell(row=row_num, column=5, value=project.teaser)
            ws.cell(row=row_num, column=6, value=project.background_objectives)
            ws.cell(row=row_num, column=7, value=project.tasks_eligibility)
            ws.cell(row=row_num, column=8, value=project.country)
            ws.cell(row=row_num, column=9, value=project.theme)
            ws.cell(row=row_num, column=10, value=project.duration)
            ws.cell(row=row_num, column=11, value=project.difficulty)
            ws.cell(row=row_num, column=12, value=project.headcount or 0)
            ws.cell(row=row_num, column=13, value=project.total_headcount)
            ws.cell(row=row_num, column=14, value=project.cover_image_url)
            ws.cell(row=row_num, column=15, value=', '.join(project.video_urls) if project.video_urls else '')
            ws.cell(row=row_num, column=16, value=', '.join(project.image_urls) if project.image_urls else '')
            ws.cell(row=row_num, column=17, value=project.application_deadline.strftime('%Y-%m-%d %H:%M:%S') if project.application_deadline else '')
            ws.cell(row=row_num, column=18, value=project.start_date.strftime('%Y-%m-%d') if project.start_date else '')
            ws.cell(row=row_num, column=19, value=project.end_date.strftime('%Y-%m-%d') if project.end_date else '')
            ws.cell(row=row_num, column=20, value='Yes' if project.is_active else 'No')
            ws.cell(row=row_num, column=21, value='Yes' if project.is_hero_highlight else 'No')
            ws.cell(row=row_num, column=22, value='Yes' if project.is_featured else 'No')
            ws.cell(row=row_num, column=23, value=project.enrolled_users.count())
            ws.cell(row=row_num, column=24, value=project.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=25, value=project.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="projects.xlsx"'
        wb.save(response)
        return response

    def export_html(self):
        queryset = self.get_queryset()
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Projects Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                h1 {{ color: #333; }}
            </style>
        </head>
        <body>
            <h1>Projects Export</h1>
            <p>Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>Total records: {queryset.count()}</p>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Project ID</th>
                        <th>KICC Project ID</th>
                        <th>Title</th>
                        <th>Teaser</th>
                        <th>Background Objectives</th>
                        <th>Tasks Eligibility</th>
                        <th>Country</th>
                        <th>Theme</th>
                        <th>Duration</th>
                        <th>Difficulty</th>
                        <th>Headcount</th>
                        <th>Total Headcount</th>
                        <th>Cover Image URL</th>
                        <th>Video URLs</th>
                        <th>Image URLs</th>
                        <th>Application Deadline</th>
                        <th>Start Date</th>
                        <th>End Date</th>
                        <th>Is Active</th>
                        <th>Is Hero Highlight</th>
                        <th>Is Featured</th>
                        <th>Enrolled Users Count</th>
                        <th>Created At</th>
                        <th>Updated At</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for project in queryset:
            html_content += f"""
                    <tr>
                        <td>{project.id}</td>
                        <td>{project.project_id}</td>
                        <td>{project.kicc_project_id or ''}</td>
                        <td>{project.title}</td>
                        <td>{project.teaser}</td>
                        <td>{project.background_objectives}</td>
                        <td>{project.tasks_eligibility}</td>
                        <td>{project.country}</td>
                        <td>{project.theme}</td>
                        <td>{project.duration}</td>
                        <td>{project.difficulty}</td>
                        <td>{project.headcount or 0}</td>
                        <td>{project.total_headcount}</td>
                        <td>{project.cover_image_url or ''}</td>
                        <td>{', '.join(project.video_urls) if project.video_urls else ''}</td>
                        <td>{', '.join(project.image_urls) if project.image_urls else ''}</td>
                        <td>{project.application_deadline.strftime('%Y-%m-%d %H:%M:%S') if project.application_deadline else ''}</td>
                        <td>{project.start_date.strftime('%Y-%m-%d') if project.start_date else ''}</td>
                        <td>{project.end_date.strftime('%Y-%m-%d') if project.end_date else ''}</td>
                        <td>{'Yes' if project.is_active else 'No'}</td>
                        <td>{'Yes' if project.is_hero_highlight else 'No'}</td>
                        <td>{'Yes' if project.is_featured else 'No'}</td>
                        <td>{project.enrolled_users.count()}</td>
                        <td>{project.created_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                        <td>{project.updated_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="projects.html"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        return context

class ProjectDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Project
    template_name = 'content_management/project_detail.html'
    context_object_name = 'project'
    queryset = Project.objects.all()
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.object
        
        # Create unified gallery items list combining all content types
        gallery_items = []
        index = 0
        
        # Add blob images
        for gallery_image in project.gallery_images.all().order_by('order'):
            gallery_items.append({
                'type': 'blob_image',
                'object': gallery_image,
                'index': index
            })
            index += 1
        
        # Add image URLs
        for image_url in project.image_urls:
            gallery_items.append({
                'type': 'image_url',
                'url': image_url,
                'index': index
            })
            index += 1
        
        # Add video URLs
        for video_url in project.video_urls:
            gallery_items.append({
                'type': 'video_url',
                'url': video_url,
                'index': index
            })
            index += 1
        
        # Paginate the unified gallery items
        from django.core.paginator import Paginator
        gallery_paginator = Paginator(gallery_items, 12)  # Show 20 items per page
        gallery_page_number = self.request.GET.get('gallery_page', 1)
        try:
            gallery_page_obj = gallery_paginator.page(gallery_page_number)
        except:
            gallery_page_obj = gallery_paginator.page(1)
        
        context['gallery_items'] = gallery_page_obj
        context['gallery_paginator'] = gallery_paginator
        
        return context

class ProjectCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Project
    template_name = 'content_management/project_form.html'
    form_class = ProjectForm
    success_url = reverse_lazy('project_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        # Process video and image URLs from form arrays
        video_urls = self.request.POST.getlist('video_urls[]')
        image_urls = self.request.POST.getlist('image_urls[]')
        
        # Filter out empty strings and save to model
        form.instance.video_urls = [url for url in video_urls if url.strip()]
        form.instance.image_urls = [url for url in image_urls if url.strip()]
        
        # Save the form first
        response = super().form_valid(form)
        
        # Handle multiple gallery image uploads using helper function
        handle_project_gallery_image_uploads(self.request, form.instance)
        
        return response

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class ProjectUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Project
    template_name = 'content_management/project_form.html'
    context_object_name = 'project'
    form_class = ProjectForm
    success_url = reverse_lazy('project_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        # Process video and image URLs from form arrays
        video_urls = self.request.POST.getlist('video_urls[]')
        image_urls = self.request.POST.getlist('image_urls[]')
        
        # Filter out empty strings and save to model
        form.instance.video_urls = [url for url in video_urls if url.strip()]
        form.instance.image_urls = [url for url in image_urls if url.strip()]
        
        # Save the form first
        response = super().form_valid(form)
        
        # Handle multiple gallery image uploads using helper function
        handle_project_gallery_image_uploads(self.request, form.instance)
        
        return response

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class ProjectDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Project
    template_name = 'content_management/project_confirm_delete.html'
    context_object_name = 'project'
    success_url = reverse_lazy('project_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

# --- NewsEvent Views ---
class NewsEventListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    paginate_by = 15
    model = NewsEvent
    template_name = 'content_management/news_event_list.html'
    context_object_name = 'news_events'
    queryset = NewsEvent.objects.all().order_by('-publish_date')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search', '')
        sort_by = self.request.GET.get('sort', '-publish_date')

        # Apply search filter
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) |
                Q(body__icontains=search_query) |
                Q(content_type__icontains=search_query)
            )

        # Apply sorting
        allowed_sort_fields = {
            'title': 'title',
            '-title': '-title',
            'content_type': 'content_type',
            '-content_type': '-content_type',
            'publish_date': 'publish_date',
            '-publish_date': '-publish_date',
            'is_published': 'is_published',
            '-is_published': '-is_published',
        }

        if sort_by in allowed_sort_fields:
            queryset = queryset.order_by(allowed_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-publish_date')  # Default sorting

        return queryset

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format:
            return self.export_data(export_format)
        return super().get(request, *args, **kwargs)

    def export_data(self, format_type):
        queryset = self.get_queryset()
        
        if format_type == 'csv':
            return self.export_csv()
        elif format_type == 'tsv':
            return self.export_tsv()
        elif format_type == 'json':
            return self.export_json()
        elif format_type == 'xlsx' and XLSX_AVAILABLE:
            return self.export_xlsx()
        elif format_type == 'html':
            return self.export_html()
        else:
            return self.export_csv()

    def export_csv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="news_events.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'News Event ID', 'Title', 'Body', 'Content Type', 'Cover Image URL', 
            'External Link', 'Video URLs', 'Image URLs', 'Publish Date', 'Is Published', 
            'Is Hero Highlight', 'Is Featured', 'Created At', 'Updated At'
        ])
        
        for item in queryset:
            writer.writerow([
                item.id,
                item.news_event_id,
                item.title,
                item.body,
                item.content_type,
                item.cover_image_url or '',
                item.external_link or '',
                ', '.join(item.video_urls) if item.video_urls else '',
                ', '.join(item.image_urls) if item.image_urls else '',
                item.publish_date.strftime('%Y-%m-%d %H:%M:%S') if item.publish_date else '',
                'Yes' if item.is_published else 'No',
                'Yes' if item.is_hero_highlight else 'No',
                'Yes' if item.is_featured else 'No',
                item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                item.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_tsv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/tab-separated-values')
        response['Content-Disposition'] = 'attachment; filename="news_events.tsv"'
        
        writer = csv.writer(response, delimiter='\t')
        writer.writerow([
            'ID', 'News Event ID', 'Title', 'Body', 'Content Type', 'Cover Image URL', 
            'External Link', 'Video URLs', 'Image URLs', 'Publish Date', 'Is Published', 
            'Is Hero Highlight', 'Is Featured', 'Created At', 'Updated At'
        ])
        
        for item in queryset:
            writer.writerow([
                item.id,
                item.news_event_id,
                item.title,
                item.body,
                item.content_type,
                item.cover_image_url or '',
                item.external_link or '',
                ', '.join(item.video_urls) if item.video_urls else '',
                ', '.join(item.image_urls) if item.image_urls else '',
                item.publish_date.strftime('%Y-%m-%d %H:%M:%S') if item.publish_date else '',
                'Yes' if item.is_published else 'No',
                'Yes' if item.is_hero_highlight else 'No',
                'Yes' if item.is_featured else 'No',
                item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                item.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_json(self):
        queryset = self.get_queryset()
        data = []
        
        for item in queryset:
            data.append({
                'id': item.id,
                'news_event_id': item.news_event_id,
                'title': item.title,
                'body': item.body,
                'content_type': item.content_type,
                'cover_image_url': item.cover_image_url,
                'external_link': item.external_link,
                'video_urls': item.video_urls,
                'image_urls': item.image_urls,
                'publish_date': item.publish_date.isoformat() if item.publish_date else None,
                'is_published': item.is_published,
                'is_hero_highlight': item.is_hero_highlight,
                'is_featured': item.is_featured,
                'created_at': item.created_at.isoformat(),
                'updated_at': item.updated_at.isoformat()
            })
        
        response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="news_events.json"'
        return response

    def export_xlsx(self):
        if not XLSX_AVAILABLE:
            return HttpResponse("XLSX export not available. Please install openpyxl.", status=400)
            
        queryset = self.get_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = "News Events"
        
        headers = [
            'ID', 'News Event ID', 'Title', 'Body', 'Content Type', 'Cover Image URL', 
            'External Link', 'Video URLs', 'Image URLs', 'Publish Date', 'Is Published', 
            'Is Hero Highlight', 'Is Featured', 'Created At', 'Updated At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, item in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=item.id)
            ws.cell(row=row_num, column=2, value=item.news_event_id)
            ws.cell(row=row_num, column=3, value=item.title)
            ws.cell(row=row_num, column=4, value=item.body)
            ws.cell(row=row_num, column=5, value=item.content_type)
            ws.cell(row=row_num, column=6, value=item.cover_image_url)
            ws.cell(row=row_num, column=7, value=item.external_link)
            ws.cell(row=row_num, column=8, value=', '.join(item.video_urls) if item.video_urls else '')
            ws.cell(row=row_num, column=9, value=', '.join(item.image_urls) if item.image_urls else '')
            ws.cell(row=row_num, column=10, value=item.publish_date.strftime('%Y-%m-%d %H:%M:%S') if item.publish_date else '')
            ws.cell(row=row_num, column=11, value='Yes' if item.is_published else 'No')
            ws.cell(row=row_num, column=12, value='Yes' if item.is_hero_highlight else 'No')
            ws.cell(row=row_num, column=13, value='Yes' if item.is_featured else 'No')
            ws.cell(row=row_num, column=14, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=15, value=item.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="news_events.xlsx"'
        wb.save(response)
        return response

    def export_html(self):
        queryset = self.get_queryset()
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>News & Events Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                h1 {{ color: #333; }}
            </style>
        </head>
        <body>
            <h1>News & Events Export</h1>
            <p>Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>Total records: {queryset.count()}</p>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>News Event ID</th>
                        <th>Title</th>
                        <th>Body</th>
                        <th>Content Type</th>
                        <th>Cover Image URL</th>
                        <th>External Link</th>
                        <th>Video URLs</th>
                        <th>Image URLs</th>
                        <th>Publish Date</th>
                        <th>Is Published</th>
                        <th>Is Hero Highlight</th>
                        <th>Is Featured</th>
                        <th>Created At</th>
                        <th>Updated At</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for item in queryset:
            html_content += f"""
                    <tr>
                        <td>{item.id}</td>
                        <td>{item.news_event_id}</td>
                        <td>{item.title}</td>
                        <td>{item.body}</td>
                        <td>{item.content_type}</td>
                        <td>{item.cover_image_url or ''}</td>
                        <td>{item.external_link or ''}</td>
                        <td>{', '.join(item.video_urls) if item.video_urls else ''}</td>
                        <td>{', '.join(item.image_urls) if item.image_urls else ''}</td>
                        <td>{item.publish_date.strftime('%Y-%m-%d %H:%M:%S') if item.publish_date else ''}</td>
                        <td>{'Yes' if item.is_published else 'No'}</td>
                        <td>{'Yes' if item.is_hero_highlight else 'No'}</td>
                        <td>{'Yes' if item.is_featured else 'No'}</td>
                        <td>{item.created_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                        <td>{item.updated_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="news_events.html"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['current_sort'] = self.request.GET.get('sort', '-publish_date')
        return context

class NewsEventDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = NewsEvent
    template_name = 'content_management/news_event_detail.html'
    context_object_name = 'news_event'
    queryset = NewsEvent.objects.all()
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        news_event = self.object
        
        # Create unified gallery items list combining all content types
        gallery_items = []
        index = 0
        
        # Add blob images
        for gallery_image in news_event.gallery_images.all().order_by('order'):
            gallery_items.append({
                'type': 'blob_image',
                'object': gallery_image,
                'index': index
            })
            index += 1
        
        # Add image URLs
        for image_url in news_event.image_urls:
            gallery_items.append({
                'type': 'image_url',
                'url': image_url,
                'index': index
            })
            index += 1
        
        # Add video URLs
        for video_url in news_event.video_urls:
            gallery_items.append({
                'type': 'video_url',
                'url': video_url,
                'index': index
            })
            index += 1
        
        # Paginate the unified gallery items
        from django.core.paginator import Paginator
        gallery_paginator = Paginator(gallery_items, 12)  # Show 20 items per page
        gallery_page_number = self.request.GET.get('gallery_page', 1)
        try:
            gallery_page_obj = gallery_paginator.page(gallery_page_number)
        except:
            gallery_page_obj = gallery_paginator.page(1)
        
        context['gallery_items'] = gallery_page_obj
        context['gallery_paginator'] = gallery_paginator
        
        return context

class NewsEventCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = NewsEvent
    template_name = 'content_management/news_event_form.html'
    form_class = NewsEventForm
    success_url = reverse_lazy('news_event_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        # Handle video and image URLs
        video_urls = self.request.POST.getlist('video_urls[]')
        image_urls = self.request.POST.getlist('image_urls[]')

        form.instance.video_urls = [url for url in video_urls if url.strip()]
        form.instance.image_urls = [url for url in image_urls if url.strip()]
        
        # Save the form first
        response = super().form_valid(form)
        
        # Handle multiple gallery image uploads using helper function
        handle_news_event_gallery_image_uploads(self.request, form.instance)
        
        return response

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class NewsEventUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = NewsEvent
    template_name = 'content_management/news_event_form.html'
    context_object_name = 'news_event'
    form_class = NewsEventForm
    success_url = reverse_lazy('news_event_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        # Handle video and image URLs
        video_urls = self.request.POST.getlist('video_urls[]')
        image_urls = self.request.POST.getlist('image_urls[]')

        form.instance.video_urls = [url for url in video_urls if url.strip()]
        form.instance.image_urls = [url for url in image_urls if url.strip()]
        
        # Save the form first
        response = super().form_valid(form)
        
        # Handle multiple gallery image uploads using helper function
        handle_news_event_gallery_image_uploads(self.request, form.instance)
        
        return response

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class NewsEventDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = NewsEvent
    template_name = 'content_management/news_event_confirm_delete.html'
    context_object_name = 'news_event'
    success_url = reverse_lazy('news_event_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

# --- SuccessStory Views ---
class SuccessStoryListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    paginate_by = 15
    model = SuccessStory
    template_name = 'content_management/success_story_list.html'
    context_object_name = 'success_stories'
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        queryset = SuccessStory.objects.select_related('related_project').all()

        # Search functionality
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) |
                Q(related_project__title__icontains=search_query) |
                Q(body__icontains=search_query)
            )

        # Sorting functionality
        sort_by = self.request.GET.get('sort', '-published_at')
        valid_sort_fields = {
            'title': 'title',
            '-title': '-title',
            'published_at': 'published_at',
            '-published_at': '-published_at',
            'related_project': 'related_project__title',
            '-related_project': '-related_project__title',
            'is_published': 'is_published',
            '-is_published': '-is_published',
        }

        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-published_at')

        return queryset

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format:
            return self.export_data(export_format)
        return super().get(request, *args, **kwargs)

    def export_data(self, format_type):
        queryset = self.get_queryset()
        
        if format_type == 'csv':
            return self.export_csv()
        elif format_type == 'tsv':
            return self.export_tsv()
        elif format_type == 'json':
            return self.export_json()
        elif format_type == 'xlsx' and XLSX_AVAILABLE:
            return self.export_xlsx()
        elif format_type == 'html':
            return self.export_html()
        else:
            return self.export_csv()

    def export_csv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="success_stories.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Success Story ID', 'Title', 'Body', 'Related Project', 'Cover Image URL', 
            'Is Hero Highlight', 'Is Featured', 'Image URLs', 'Video URLs', 'Beneficiaries', 
            'Total Hours Contributed', 'Is Published', 'Published At', 'Created At', 'Updated At'
        ])
        
        for story in queryset:
            writer.writerow([
                story.id,
                story.success_story_id,
                story.title,
                story.body,
                story.related_project.title if story.related_project else '',
                story.cover_image_url or '',
                'Yes' if story.is_hero_highlight else 'No',
                'Yes' if story.is_featured else 'No',
                ', '.join(story.image_urls) if story.image_urls else '',
                ', '.join(story.video_urls) if story.video_urls else '',
                story.beneficiaries or '',
                story.total_hours_contributed or '',
                'Yes' if story.is_published else 'No',
                story.published_at.strftime('%Y-%m-%d %H:%M:%S') if story.published_at else '',
                story.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                story.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_tsv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/tab-separated-values')
        response['Content-Disposition'] = 'attachment; filename="success_stories.tsv"'
        
        writer = csv.writer(response, delimiter='\t')
        writer.writerow([
            'ID', 'Success Story ID', 'Title', 'Body', 'Related Project', 'Cover Image URL', 
            'Is Hero Highlight', 'Is Featured', 'Image URLs', 'Video URLs', 'Beneficiaries', 
            'Total Hours Contributed', 'Is Published', 'Published At', 'Created At', 'Updated At'
        ])
        
        for story in queryset:
            writer.writerow([
                story.id,
                story.success_story_id,
                story.title,
                story.body,
                story.related_project.title if story.related_project else '',
                story.cover_image_url or '',
                'Yes' if story.is_hero_highlight else 'No',
                'Yes' if story.is_featured else 'No',
                ', '.join(story.image_urls) if story.image_urls else '',
                ', '.join(story.video_urls) if story.video_urls else '',
                story.beneficiaries or '',
                story.total_hours_contributed or '',
                'Yes' if story.is_published else 'No',
                story.published_at.strftime('%Y-%m-%d %H:%M:%S') if story.published_at else '',
                story.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                story.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_json(self):
        queryset = self.get_queryset()
        data = []
        
        for story in queryset:
            data.append({
                'id': story.id,
                'success_story_id': story.success_story_id,
                'title': story.title,
                'body': story.body,
                'related_project': story.related_project.title if story.related_project else None,
                'cover_image_url': story.cover_image_url,
                'is_hero_highlight': story.is_hero_highlight,
                'is_featured': story.is_featured,
                'image_urls': story.image_urls,
                'video_urls': story.video_urls,
                'beneficiaries': story.beneficiaries,
                'total_hours_contributed': story.total_hours_contributed,
                'is_published': story.is_published,
                'published_at': story.published_at.isoformat() if story.published_at else None,
                'created_at': story.created_at.isoformat(),
                'updated_at': story.updated_at.isoformat()
            })
        
        response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="success_stories.json"'
        return response

    def export_xlsx(self):
        if not XLSX_AVAILABLE:
            return HttpResponse("XLSX export not available. Please install openpyxl.", status=400)
            
        queryset = self.get_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = "Success Stories"
        
        headers = [
            'ID', 'Success Story ID', 'Title', 'Body', 'Related Project', 'Cover Image URL', 
            'Is Hero Highlight', 'Is Featured', 'Image URLs', 'Video URLs', 'Beneficiaries', 
            'Total Hours Contributed', 'Is Published', 'Published At', 'Created At', 'Updated At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, story in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=story.id)
            ws.cell(row=row_num, column=2, value=story.success_story_id)
            ws.cell(row=row_num, column=3, value=story.title)
            ws.cell(row=row_num, column=4, value=story.body)
            ws.cell(row=row_num, column=5, value=story.related_project.title if story.related_project else '')
            ws.cell(row=row_num, column=6, value=story.cover_image_url)
            ws.cell(row=row_num, column=7, value='Yes' if story.is_hero_highlight else 'No')
            ws.cell(row=row_num, column=8, value='Yes' if story.is_featured else 'No')
            ws.cell(row=row_num, column=9, value=', '.join(story.image_urls) if story.image_urls else '')
            ws.cell(row=row_num, column=10, value=', '.join(story.video_urls) if story.video_urls else '')
            ws.cell(row=row_num, column=11, value=story.beneficiaries)
            ws.cell(row=row_num, column=12, value=story.total_hours_contributed)
            ws.cell(row=row_num, column=13, value='Yes' if story.is_published else 'No')
            ws.cell(row=row_num, column=14, value=story.published_at.strftime('%Y-%m-%d %H:%M:%S') if story.published_at else '')
            ws.cell(row=row_num, column=15, value=story.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=16, value=story.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="success_stories.xlsx"'
        wb.save(response)
        return response

    def export_html(self):
        queryset = self.get_queryset()
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Success Stories Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                h1 {{ color: #333; }}
            </style>
        </head>
        <body>
            <h1>Success Stories Export</h1>
            <p>Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>Total records: {queryset.count()}</p>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Success Story ID</th>
                        <th>Title</th>
                        <th>Body</th>
                        <th>Related Project</th>
                        <th>Cover Image URL</th>
                        <th>Is Hero Highlight</th>
                        <th>Is Featured</th>
                        <th>Image URLs</th>
                        <th>Video URLs</th>
                        <th>Beneficiaries</th>
                        <th>Total Hours Contributed</th>
                        <th>Is Published</th>
                        <th>Published At</th>
                        <th>Created At</th>
                        <th>Updated At</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for story in queryset:
            html_content += f"""
                    <tr>
                        <td>{story.id}</td>
                        <td>{story.success_story_id}</td>
                        <td>{story.title}</td>
                        <td>{story.body}</td>
                        <td>{story.related_project.title if story.related_project else ''}</td>
                        <td>{story.cover_image_url or ''}</td>
                        <td>{'Yes' if story.is_hero_highlight else 'No'}</td>
                        <td>{'Yes' if story.is_featured else 'No'}</td>
                        <td>{', '.join(story.image_urls) if story.image_urls else ''}</td>
                        <td>{', '.join(story.video_urls) if story.video_urls else ''}</td>
                        <td>{story.beneficiaries or ''}</td>
                        <td>{story.total_hours_contributed or ''}</td>
                        <td>{'Yes' if story.is_published else 'No'}</td>
                        <td>{story.published_at.strftime('%Y-%m-%d %H:%M:%S') if story.published_at else ''}</td>
                        <td>{story.created_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                        <td>{story.updated_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="success_stories.html"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['current_sort'] = self.request.GET.get('sort', '-published_at')
        return context

class SuccessStoryDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = SuccessStory
    template_name = 'content_management/success_story_detail.html'
    context_object_name = 'success_story'
    queryset = SuccessStory.objects.all()
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        success_story = self.object
        
        # Create unified gallery items list combining all content types
        gallery_items = []
        index = 0
        
        # Add blob images
        for gallery_image in success_story.gallery_images.all().order_by('order'):
            gallery_items.append({
                'type': 'blob_image',
                'object': gallery_image,
                'index': index
            })
            index += 1
        
        # Add image URLs
        for image_url in success_story.image_urls:
            gallery_items.append({
                'type': 'image_url',
                'url': image_url,
                'index': index
            })
            index += 1
        
        # Add video URLs
        for video_url in success_story.video_urls:
            gallery_items.append({
                'type': 'video_url',
                'url': video_url,
                'index': index
            })
            index += 1
        
        # Paginate the unified gallery items
        from django.core.paginator import Paginator
        gallery_paginator = Paginator(gallery_items, 12)  # Show 20 items per page
        gallery_page_number = self.request.GET.get('gallery_page', 1)
        try:
            gallery_page_obj = gallery_paginator.page(gallery_page_number)
        except:
            gallery_page_obj = gallery_paginator.page(1)
        
        context['gallery_items'] = gallery_page_obj
        context['gallery_paginator'] = gallery_paginator
        
        return context

class SuccessStoryCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = SuccessStory
    template_name = 'content_management/success_story_form.html'
    form_class = SuccessStoryForm
    success_url = reverse_lazy('success_story_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        # Process multiple image and video URLs from form arrays
        image_urls = self.request.POST.getlist('image_urls[]')
        video_urls = self.request.POST.getlist('video_urls[]')

        # Filter out empty strings and save to model
        form.instance.image_urls = [url for url in image_urls if url.strip()]
        form.instance.video_urls = [url for url in video_urls if url.strip()]

        # Save the form first
        response = super().form_valid(form)
        
        # Handle multiple gallery image uploads using helper function
        handle_gallery_image_uploads(self.request, form.instance)
        
        return response

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class SuccessStoryUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = SuccessStory
    template_name = 'content_management/success_story_form.html'
    context_object_name = 'success_story' # Ensure the object is available as 'success_story' in the template
    form_class = SuccessStoryForm
    success_url = reverse_lazy('success_story_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_valid(self, form):
        # Process multiple image and video URLs from form arrays
        image_urls = self.request.POST.getlist('image_urls[]')
        video_urls = self.request.POST.getlist('video_urls[]')

        # Filter out empty strings and save to model
        form.instance.image_urls = [url for url in image_urls if url.strip()]
        form.instance.video_urls = [url for url in video_urls if url.strip()]

        # Save the form first
        response = super().form_valid(form)
        
        # Handle multiple gallery image uploads using helper function
        handle_gallery_image_uploads(self.request, form.instance)
        
        return response

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class SuccessStoryDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = SuccessStory
    template_name = 'content_management/success_story_confirm_delete.html'
    context_object_name = 'success_story'
    success_url = reverse_lazy('success_story_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

# --- FAQ Views ---
class FAQListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    paginate_by = 15
    model = FAQ
    template_name = 'content_management/faq_list.html'
    context_object_name = 'faqs'
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def get_queryset(self):
        queryset = FAQ.objects.annotate(
            total_votes_cnt=F('thumbs_up') + F('thumbs_down')
        ).annotate(
            usefulness=Case(
                When(total_votes_cnt__gt=0, then=ExpressionWrapper(
                    F('thumbs_up') * 100.0 / F('total_votes_cnt'),
                    output_field=FloatField()
                )),
                default=Value(0.0),
                output_field=FloatField()
            )
        )

        # Search functionality
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(question__icontains=search_query) |
                Q(answer__icontains=search_query)
            )

        # Sorting functionality
        sort_by = self.request.GET.get('sort', 'order')
        valid_sort_fields = {
            'order': 'order',
            '-order': '-order',
            'question': 'question',
            '-question': '-question',
            'is_schema_ready': 'is_schema_ready',
            '-is_schema_ready': '-is_schema_ready',
            'thumbs_up': 'thumbs_up',
            '-thumbs_up': '-thumbs_up',
            'thumbs_down': 'thumbs_down',
            '-thumbs_down': '-thumbs_down',
            'usefulness': 'usefulness',
            '-usefulness': '-usefulness',
        }

        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('order')

        return queryset

    def get(self, request, *args, **kwargs):
        export_format = request.GET.get('export')
        if export_format:
            return self.export_data(export_format)
        return super().get(request, *args, **kwargs)

    def export_data(self, format_type):
        queryset = self.get_queryset()
        
        if format_type == 'csv':
            return self.export_csv()
        elif format_type == 'tsv':
            return self.export_tsv()
        elif format_type == 'json':
            return self.export_json()
        elif format_type == 'xlsx' and XLSX_AVAILABLE:
            return self.export_xlsx()
        elif format_type == 'html':
            return self.export_html()
        else:
            return self.export_csv()

    def export_csv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="faqs.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'ID', 'FAQ ID', 'Question', 'Answer', 'Order', 'Is Schema Ready', 
            'Thumbs Up', 'Thumbs Down', 'Total Votes', 'Helpfulness Ratio %', 'Created At', 'Updated At'
        ])
        
        for faq in queryset:
            writer.writerow([
                faq.id,
                faq.faq_id,
                faq.question,
                faq.answer,
                faq.order,
                'Yes' if faq.is_schema_ready else 'No',
                faq.thumbs_up,
                faq.thumbs_down,
                faq.total_votes,
                f"{faq.helpfulness_ratio}%" if faq.total_votes > 0 else '0%',
                faq.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                faq.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_tsv(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='text/tab-separated-values')
        response['Content-Disposition'] = 'attachment; filename="faqs.tsv"'
        
        writer = csv.writer(response, delimiter='\t')
        writer.writerow([
            'ID', 'FAQ ID', 'Question', 'Answer', 'Order', 'Is Schema Ready', 
            'Thumbs Up', 'Thumbs Down', 'Total Votes', 'Helpfulness Ratio %', 'Created At', 'Updated At'
        ])
        
        for faq in queryset:
            writer.writerow([
                faq.id,
                faq.faq_id,
                faq.question,
                faq.answer,
                faq.order,
                'Yes' if faq.is_schema_ready else 'No',
                faq.thumbs_up,
                faq.thumbs_down,
                faq.total_votes,
                f"{faq.helpfulness_ratio}%" if faq.total_votes > 0 else '0%',
                faq.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                faq.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response

    def export_json(self):
        queryset = self.get_queryset()
        data = []
        
        for faq in queryset:
            data.append({
                'id': faq.id,
                'faq_id': faq.faq_id,
                'question': faq.question,
                'answer': faq.answer,
                'order': faq.order,
                'is_schema_ready': faq.is_schema_ready,
                'thumbs_up': faq.thumbs_up,
                'thumbs_down': faq.thumbs_down,
                'total_votes': faq.total_votes,
                'helpfulness_ratio': faq.helpfulness_ratio,
                'created_at': faq.created_at.isoformat(),
                'updated_at': faq.updated_at.isoformat()
            })
        
        response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="faqs.json"'
        return response

    def export_xlsx(self):
        if not XLSX_AVAILABLE:
            return HttpResponse("XLSX export not available. Please install openpyxl.", status=400)
            
        queryset = self.get_queryset()
        wb = Workbook()
        ws = wb.active
        ws.title = "FAQs"
        
        headers = [
            'ID', 'FAQ ID', 'Question', 'Answer', 'Order', 'Is Schema Ready', 
            'Thumbs Up', 'Thumbs Down', 'Total Votes', 'Helpfulness Ratio %', 'Created At', 'Updated At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_num, faq in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=faq.id)
            ws.cell(row=row_num, column=2, value=faq.faq_id)
            ws.cell(row=row_num, column=3, value=faq.question)
            ws.cell(row=row_num, column=4, value=faq.answer)
            ws.cell(row=row_num, column=5, value=faq.order)
            ws.cell(row=row_num, column=6, value='Yes' if faq.is_schema_ready else 'No')
            ws.cell(row=row_num, column=7, value=faq.thumbs_up)
            ws.cell(row=row_num, column=8, value=faq.thumbs_down)
            ws.cell(row=row_num, column=9, value=faq.total_votes)
            ws.cell(row=row_num, column=10, value=f"{faq.helpfulness_ratio}%" if faq.total_votes > 0 else '0%')
            ws.cell(row=row_num, column=11, value=faq.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=12, value=faq.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="faqs.xlsx"'
        wb.save(response)
        return response

    def export_html(self):
        queryset = self.get_queryset()
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>FAQs Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                h1 {{ color: #333; }}
            </style>
        </head>
        <body>
            <h1>FAQs Export</h1>
            <p>Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>Total records: {queryset.count()}</p>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>FAQ ID</th>
                        <th>Question</th>
                        <th>Answer</th>
                        <th>Order</th>
                        <th>Is Schema Ready</th>
                        <th>Thumbs Up</th>
                        <th>Thumbs Down</th>
                        <th>Total Votes</th>
                        <th>Helpfulness Ratio %</th>
                        <th>Created At</th>
                        <th>Updated At</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for faq in queryset:
            html_content += f"""
                    <tr>
                        <td>{faq.id}</td>
                        <td>{faq.faq_id}</td>
                        <td>{faq.question}</td>
                        <td>{faq.answer}</td>
                        <td>{faq.order}</td>
                        <td>{'Yes' if faq.is_schema_ready else 'No'}</td>
                        <td>{faq.thumbs_up}</td>
                        <td>{faq.thumbs_down}</td>
                        <td>{faq.total_votes}</td>
                        <td>{f"{faq.helpfulness_ratio}%" if faq.total_votes > 0 else '0%'}</td>
                        <td>{faq.created_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                        <td>{faq.updated_at.strftime('%Y-%m-%d %H:%M:%S')}</td>
                    </tr>
            """
        
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
        
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="faqs.html"'
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['current_sort'] = self.request.GET.get('sort', 'order')
        return context

class FAQDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = FAQ
    template_name = 'content_management/faq_detail.html'
    context_object_name = 'faq'
    queryset = FAQ.objects.all()
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

class FAQCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = FAQ
    template_name = 'content_management/faq_form.html'
    form_class = FAQForm
    success_url = reverse_lazy('faq_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class FAQUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = FAQ
    template_name = 'content_management/faq_form.html'
    form_class = FAQForm
    success_url = reverse_lazy('faq_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff

    def form_invalid(self, form):
        logger.error(f"Form has errors: {form.errors.as_json()}")
        return super().form_invalid(form)

class FAQDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = FAQ
    template_name = 'content_management/faq_confirm_delete.html'
    context_object_name = 'faq'
    success_url = reverse_lazy('faq_list')
    login_url = '/login/'
    redirect_field_name = 'next'

    def test_func(self):
        return self.request.user.is_staff
